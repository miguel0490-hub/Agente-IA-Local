import os
import markdown
import io
import datetime
import re
import html
from src.core.config import CARPETA_IMAGENES

# Imports pesados movidos al interior de los métodos para Lazy Loading

# Compatibilidad legacy: exponer disponibilidad/config de pdfkit a nivel módulo.
HAS_PDFKIT = False
PDFKIT_CONFIG = None
try:
    import pdfkit
    import platform

    _default_wk = (
        r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        if platform.system() == "Windows"
        else "/usr/bin/wkhtmltopdf"
    )
    _wkhtmltopdf_path = os.getenv("WKHTMLTOPDF_PATH", _default_wk)
    if _wkhtmltopdf_path and os.path.exists(_wkhtmltopdf_path):
        PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf=_wkhtmltopdf_path)
    HAS_PDFKIT = True
except ImportError:
    HAS_PDFKIT = False

class FileFactory:
    """Fábrica de archivos multiformato invocable por el LLM."""
    def __init__(self, output_dir=CARPETA_IMAGENES):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def execute_tool(self, tool_data: dict) -> str:
        """
        Ejecuta la herramienta de creación o edición basándose en el JSON.
        Retorna la ruta absoluta del archivo resultante o None si falló.
        """
        import os
        from pathlib import Path
        import datetime

        raw_filename = tool_data.get("filename", f"file_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.txt")
        # Sanitización estricta: extraer solo el nombre base, eliminando rutas relativas (../)
        safe_filename = Path(raw_filename).name
        if not safe_filename or safe_filename.startswith('.'):
            safe_filename = f"file_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
        filepath = os.path.join(self.output_dir, safe_filename)
        
        action = tool_data.get("action")
        content = tool_data.get("content", "")
        
        try:
            if action == "create_file":
                if safe_filename.lower().endswith(".pdf"):
                    return self._create_pdf(filepath, content)
                elif safe_filename.lower().endswith((".xlsx", ".xls")):
                    return self._create_excel(filepath, content)
                elif safe_filename.lower().endswith(".html"):
                    return self._create_text(filepath, content)
                else:
                    return self._create_text(filepath, content)
            elif action == "edit_file":
                search = tool_data.get("search", "")
                replace = tool_data.get("replace", "")
                return self._edit_text(filepath, search, replace)
            return None
        except Exception as e:
            print(f"Error ejecutando herramienta {action}: {e}")
            return None

    def _create_text(self, filepath, content):
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        return filepath

    def _create_pdf(self, filepath, content):
        """
        Crea un PDF a partir del contenido recibido.
        Pipeline de prioridad:
          1. Si el contenido es HTML → pdfkit (Print CSS nativo)
          2. Si el contenido es HTML pero no hay pdfkit → guarda como .html descargable
          3. Si el contenido es Markdown → ReportLab (fallback legacy)
          4. Sin ninguna librería → guarda como .md
        """
        # Segunda línea de defensa: si aún llegan \\n literales (LLM no escapó correctamente),
        # los convertimos aquí antes de cualquier detección o escritura.
        if "\\n" in content and "\n" not in content:
            content = content.replace("\\n", "\n").replace("\\t", "\t").replace('\\"', '"')

        # Detectar si el contenido es HTML buscando en TODO el string (no solo el inicio),
        # por si el LLM antepuso texto antes del <!DOCTYPE>.
        content_lower = content.lower()
        content_is_html = (
            "<!doctype html" in content_lower
            or "<html" in content_lower
            or ("<head>" in content_lower and "<body>" in content_lower)
        )
        # Si es HTML, recortamos cualquier texto previo al <!DOCTYPE para entregarlo limpio
        if content_is_html:
            for marker in ["<!doctype html", "<!DOCTYPE html", "<html", "<HTML"]:
                idx = content.find(marker)
                if idx > 0:
                    content = content[idx:]
                    break

        # ── Rama 1 & 2: Contenido HTML ───────────────────────────────────────
        if content_is_html:
            content = self._enforce_pdf_layout_guardrails(content)
            HAS_PDFKIT = False
            PDFKIT_CONFIG = None
            try:
                import pdfkit
                import platform
                _default_wk = (r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
                               if platform.system() == "Windows"
                               else "/usr/bin/wkhtmltopdf")
                WKHTMLTOPDF_PATH = os.getenv("WKHTMLTOPDF_PATH", _default_wk)
                if WKHTMLTOPDF_PATH and os.path.exists(WKHTMLTOPDF_PATH):
                    PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
                HAS_PDFKIT = True
            except ImportError:
                pass

            if HAS_PDFKIT:
                # Estrategia: escribir el HTML a un fichero temporal y convertir desde
                # disco con from_file(). Esto elimina todos los problemas de encoding
                # y caracteres especiales que from_string() tiene con HTML complejo.
                tmp_html_path = filepath.replace(".pdf", "_tmp_source.html")
                try:
                    # 1. Escribir HTML limpio al disco
                    with open(tmp_html_path, "w", encoding="utf-8") as f:
                        f.write(content)

                    # 2. Convertir desde fichero (mucho más robusto que from_string)
                    options = {
                        "page-size":    "A4",
                        "margin-top":   "2.5cm",
                        "margin-right": "2.5cm",
                        "margin-bottom":"2.5cm",
                        "margin-left":  "2.5cm",
                        "encoding":     "UTF-8",
                        "enable-local-file-access": "",
                        "quiet":        "",
                    }
                    pdfkit.from_file(tmp_html_path, filepath, options=options, configuration=PDFKIT_CONFIG)

                    # 3. Verificar que el PDF se generó y tiene contenido real
                    if os.path.exists(filepath) and os.path.getsize(filepath) > 1024:
                        return filepath
                    else:
                        raise RuntimeError(f"PDF generado está vacío o es inválido ({os.path.getsize(filepath)} bytes)")

                except Exception as pdfkit_err:
                    import traceback
                    print(f"[FileFactory][ERROR] pdfkit.from_file falló:")
                    print(traceback.format_exc())
                    # Intentar from_string como segunda opción antes del fallback HTML
                    try:
                        pdfkit.from_string(content, filepath, options=options, configuration=PDFKIT_CONFIG)
                        if os.path.exists(filepath) and os.path.getsize(filepath) > 1024:
                            return filepath
                    except Exception as fs_err:
                        print(f"[FileFactory][ERROR] pdfkit.from_string también falló: {fs_err}")
                finally:
                    # Limpiar el HTML temporal (éxito o error)
                    if os.path.exists(tmp_html_path):
                        os.remove(tmp_html_path)

            # Fallback robusto: convertir HTML a texto y generar PDF con ReportLab.
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet

                text_content = self._html_to_text(content)
                doc = SimpleDocTemplate(filepath, pagesize=letter)
                styles = getSampleStyleSheet()
                flowables = []
                for line in text_content.split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith("#"):
                        flowables.append(Paragraph(f"<b>{line.lstrip('#').strip()}</b>", styles["Heading1"]))
                    else:
                        flowables.append(Paragraph(line, styles["Normal"]))
                    flowables.append(Spacer(1, 10))
                doc.build(flowables)
                if os.path.exists(filepath) and os.path.getsize(filepath) > 512:
                    return filepath
            except Exception as fallback_err:
                print(f"[FileFactory] Fallback HTML->PDF con ReportLab falló: {fallback_err}")

            # Último recurso: guardar HTML descargable
            html_filepath = filepath.replace(".pdf", ".html")
            return self._create_text(html_filepath, content)

        # ── Rama 3: Contenido Markdown → ReportLab ───────────────────────────
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            HAS_REPORTLAB = True
        except ImportError:
            HAS_REPORTLAB = False

        if HAS_REPORTLAB:
            try:
                doc = SimpleDocTemplate(filepath, pagesize=letter)
                styles = getSampleStyleSheet()
                flowables = []
                for line in content.split("\n"):
                    line = line.strip()
                    if line:
                        if line.startswith("#"):
                            clean_line = line.lstrip("#").strip()
                            flowables.append(Paragraph(f"<b>{clean_line}</b>", styles["Heading1"]))
                        else:
                            flowables.append(Paragraph(line, styles["Normal"]))
                        flowables.append(Spacer(1, 10))
                doc.build(flowables)
                return filepath
            except Exception as rl_err:
                print(f"[FileFactory] ReportLab falló: {rl_err}. Fallback a .md")

        # ── Rama 4: Sin librerías → guardar como Markdown plano ───────────────
        md_filepath = filepath.replace(".pdf", ".md")
        return self._create_text(md_filepath, content)

    def _create_excel(self, filepath, content):
        try:
            import pandas as pd
            HAS_PANDAS = True
        except ImportError:
            HAS_PANDAS = False

        if not HAS_PANDAS:
            filepath = filepath.replace('.xlsx', '.csv')
            return self._create_text(filepath, content)

        try:
            tables = self._extract_markdown_tables(content)

            if not tables:
                # Fallback: intentar leer como CSV puro
                df = pd.read_csv(io.StringIO(content))
                tables = [("Datos", [], df)]

            # Escribir todas las tablas en hojas separadas
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, col_alignments, df in tables:
                    safe_name = sheet_name[:31]  # Excel limita a 31 chars por hoja
                    df.to_excel(writer, index=False, sheet_name=safe_name)

            # Formateo Premium con openpyxl
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = load_workbook(filepath)

                HEADER_FILL       = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                TOTAL_FILL        = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                HEADER_FONT       = Font(name="Calibri", color="00F2FE", bold=True, size=11)
                TOTAL_FONT        = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
                DATA_FONT         = Font(name="Calibri", color="E2E8F0", size=10)
                THIN_BORDER_SIDE  = Side(style="thin", color="334155")
                CELL_BORDER       = Border(
                    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
                    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE
                )

                align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
                align_center = Alignment(horizontal="center", vertical="center")
                align_right  = Alignment(horizontal="right",  vertical="center")

                # Indicadores que determinan si una columna debe ir a la derecha
                RIGHT_KEYWORDS = {"$", "€", "%", "importe", "total", "precio",
                                  "monto", "coste", "costo", "ingreso", "margen",
                                  "roi", "beneficio", "puntuación", "promedio", "valor"}

                for ws, (_, col_alignments, df) in zip(wb.worksheets, tables):
                    ws.freeze_panes = "A2"
                    ws.sheet_view.showGridLines = False

                    # ── Estilo de cabeceras ──────────────────────────────────────
                    for col_idx, cell in enumerate(ws[1], 1):
                        cell.fill       = HEADER_FILL
                        cell.font       = HEADER_FONT
                        cell.alignment  = align_center
                        cell.border     = CELL_BORDER
                        ws.row_dimensions[1].height = 22

                    # ── Estilo de filas de datos ──────────────────────────────────
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        is_total_row = False
                        for col_idx, cell in enumerate(row, 1):
                            raw_val = str(cell.value or "").strip().upper()

                            # Detectar fila TOTAL
                            if col_idx == 1 and raw_val in ("TOTAL", "**TOTAL**"):
                                is_total_row = True

                            # Determinar alineación: usar la del Markdown si existe, sino inferir
                            md_align = col_alignments[col_idx - 1] if col_idx <= len(col_alignments) else "left"
                            header_name = str(df.columns[col_idx - 1]).lower() if col_idx <= len(df.columns) else ""
                            is_numeric_col = any(kw in header_name for kw in RIGHT_KEYWORDS)
                            if md_align == "right" or is_numeric_col:
                                cell.alignment = align_right
                            elif md_align == "center":
                                cell.alignment = align_center
                            else:
                                cell.alignment = align_left

                            cell.border = CELL_BORDER

                            # Alternar color de fila
                            if is_total_row:
                                cell.fill = TOTAL_FILL
                                cell.font = TOTAL_FONT
                            elif row_idx % 2 == 0:
                                cell.fill = PatternFill(start_color="1A2333", end_color="1A2333", fill_type="solid")
                                cell.font = DATA_FONT
                            else:
                                cell.fill = PatternFill(start_color="131C28", end_color="131C28", fill_type="solid")
                                cell.font = DATA_FONT

                        ws.row_dimensions[row_idx].height = 18

                    # ── Auto-ajuste de ancho de columnas ─────────────────────────
                    for col_idx, col in enumerate(ws.columns, 1):
                        col_letter = get_column_letter(col_idx)
                        max_len = max(
                            (len(str(cell.value or "")) for cell in col),
                            default=10
                        )
                        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 16), 60)

                wb.save(filepath)
            except Exception as fmt_err:
                print(f"[FileFactory] Formato premium omitido: {fmt_err}")

            return filepath

        except Exception as e:
            print(f"[FileFactory] Error convirtiendo a Excel. Fallback CSV: {e}")
            filepath = filepath.replace('.xlsx', '.csv')
            with open(filepath, "w", encoding="utf-8-sig") as f:
                f.write(content)
            return filepath

    # ─────────────────────────────────────────────────────────────────────────
    # Utilidades privadas
    # ─────────────────────────────────────────────────────────────────────────

    def _extract_markdown_tables(self, content: str) -> list:
        """
        Extrae todas las tablas Markdown de un bloque de texto.
        Maneja: negrita en encabezados, filas de alineación (:---:, ---:, :---),
        múltiples tablas separadas por texto, y limpieza de caracteres de formato.
        Retorna una lista de tuplas: (nombre_hoja, [alineaciones], DataFrame).
        """
        import re

        tables = []
        # Dividir el contenido en bloques y encontrar tablas junto a sus títulos
        title_pattern = re.compile(r"(?:#{1,6}\s*(.+?)\n)((?:[^\n]*\|[^\n]*\n)+)", re.MULTILINE)
        bare_table_pattern = re.compile(r"((?:[^\n]*\|[^\n]*\n)+)", re.MULTILINE)

        found_spans = []

        for match in title_pattern.finditer(content):
            title = re.sub(r"\*+", "", match.group(1)).strip()
            table_block = match.group(2)
            result = self._parse_single_markdown_table(table_block)
            if result:
                col_alignments, df = result
                tables.append((title or f"Tabla {len(tables)+1}", col_alignments, df))
            found_spans.append(match.span())

        # Extraer tablas sin título que no estén ya incluidas
        for match in bare_table_pattern.finditer(content):
            start, end = match.span()
            already_captured = any(s <= start and end <= e for s, e in found_spans)
            if not already_captured:
                result = self._parse_single_markdown_table(match.group(1))
                if result:
                    col_alignments, df = result
                    tables.append((f"Tabla {len(tables)+1}", col_alignments, df))

        return tables

    def _parse_single_markdown_table(self, block: str):
        """
        Parsea un único bloque de tabla Markdown.
        Retorna (lista_alineaciones, DataFrame) o None si el bloque no es válido.
        """
        import pandas as pd
        import re

        lines = [l.strip() for l in block.strip().splitlines() if l.strip().startswith("|")]
        if len(lines) < 2:
            return None

        def clean_cell(cell: str) -> str:
            """Elimina negrita, cursiva y espacios sobrantes."""
            cell = cell.strip()
            cell = re.sub(r"\*+", "", cell)   # quita ** y *
            cell = re.sub(r"_+", "", cell)    # quita __
            return cell.strip()

        def detect_alignment(sep: str) -> str:
            sep = sep.strip()
            if sep.startswith(":") and sep.endswith(":"):
                return "center"
            if sep.endswith(":"):
                return "right"
            return "left"

        # Detectar la fila de alineación (siempre contiene ---)
        sep_idx = next((i for i, l in enumerate(lines) if re.search(r":?-{2,}:?", l)), None)
        if sep_idx is None:
            return None

        header_line = lines[0]
        sep_line    = lines[sep_idx]

        headers = [clean_cell(h) for h in header_line.strip("|").split("|")]
        col_alignments = [detect_alignment(s) for s in sep_line.strip("|").split("|")]

        # Alinear número de columnas por si hay desajuste
        num_cols = max(len(headers), len(col_alignments))
        while len(headers) < num_cols:       headers.append("")
        while len(col_alignments) < num_cols: col_alignments.append("left")

        data_lines = [l for i, l in enumerate(lines) if i != 0 and i != sep_idx]
        rows = []
        for line in data_lines:
            cells = [clean_cell(c) for c in line.strip("|").split("|")]
            while len(cells) < num_cols:
                cells.append("")
            rows.append(cells[:num_cols])

        if not rows:
            return None

        df = pd.DataFrame(rows, columns=headers[:num_cols])
        return col_alignments[:num_cols], df
            
    def _edit_text(self, filepath, search, replace):
        if not os.path.exists(filepath):
            return None
        with open(filepath, "r", encoding="utf-8") as f:
            data = f.read()
        
        data = data.replace(search, replace)
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(data)
        return filepath

    def _html_to_text(self, html_content: str) -> str:
        """Convierte HTML simple a texto legible para fallback PDF."""
        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", "", html_content)
        text = re.sub(r"(?i)</(p|div|section|article|h1|h2|h3|h4|h5|h6|li|tr|br)>", "\n", text)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        text = html.unescape(text)
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n\s+\n", "\n\n", text)
        return text.strip()

    def _enforce_pdf_layout_guardrails(self, html_content: str) -> str:
        """
        Inyecta reglas CSS de impresión para evitar títulos huérfanos y cortes bruscos.
        Se aplica sobre HTML generado por el LLM antes de pasarlo a pdfkit.
        """
        html_content = self._apply_corporate_print_template(html_content)
        html_content = self._group_headings_with_following_block(html_content)
        guardrail_css = """
<style id="superagente-pdf-guardrails">
@page {
  size: A4;
  margin: 2.4cm 2.2cm 2.2cm 2.2cm;
}
body {
  font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
  font-size: 11.2pt;
  line-height: 1.45;
  color: #1f2937;
  margin: 0;
  padding: 1.2cm 0 1.4cm 0;
}
h1, h2, h3, h4, h5, h6 {
  page-break-after: avoid !important;
  break-after: avoid-page !important;
  page-break-inside: avoid !important;
  break-inside: avoid !important;
  orphans: 3 !important;
  widows: 3 !important;
  margin-top: 14px !important;
  margin-bottom: 8px !important;
  line-height: 1.25 !important;
}
p {
  margin: 0 0 9px 0 !important;
  page-break-inside: auto !important;
  break-inside: auto !important;
  text-align: justify !important;
}
li {
  margin-bottom: 4px !important;
  page-break-inside: auto !important;
  break-inside: auto !important;
}
table, figure, blockquote {
  page-break-inside: avoid !important;
  break-inside: avoid-page !important;
  orphans: 3 !important;
  widows: 3 !important;
}
section, article, .section, .bloque {
  page-break-inside: avoid !important;
  break-inside: avoid-page !important;
}
.sa-keep-with-next {
  page-break-inside: avoid !important;
  break-inside: avoid-page !important;
  margin-bottom: 6px !important;
}
.sa-corp-header {
  position: fixed;
  top: -1.2cm;
  left: 0;
  right: 0;
  font-size: 9pt;
  color: #64748b;
  border-bottom: 1px solid #e2e8f0;
  padding-bottom: 4px;
}
.sa-corp-footer {
  position: fixed;
  bottom: -1.1cm;
  left: 0;
  right: 0;
  font-size: 9pt;
  color: #64748b;
  border-top: 1px solid #e2e8f0;
  padding-top: 4px;
}
.sa-corp-footer .sa-page-number::before {
  content: counter(page);
}
</style>
"""
        if "superagente-pdf-guardrails" in html_content:
            return html_content
        if "</head>" in html_content.lower():
            return re.sub(r"(?i)</head>", f"{guardrail_css}\n</head>", html_content, count=1)
        return f"{guardrail_css}\n{html_content}"

    def _group_headings_with_following_block(self, html_content: str) -> str:
        """
        Agrupa encabezado + primer bloque de contenido para evitar encabezados huérfanos.
        """
        pattern = re.compile(
            r"(?is)"
            r"(<h[1-6][^>]*>.*?</h[1-6]>)"
            r"(\s*(?:<p[^>]*>.*?</p>|<ul[^>]*>.*?</ul>|<ol[^>]*>.*?</ol>|<table[^>]*>.*?</table>|<div[^>]*>.*?</div>|<blockquote[^>]*>.*?</blockquote>))"
        )
        return pattern.sub(r'<div class="sa-keep-with-next">\1\2</div>', html_content)

    def _apply_corporate_print_template(self, html_content: str) -> str:
        """Inyecta cabecera y pie corporativos consistentes para salida PDF."""
        if "sa-corp-header" in html_content and "sa-corp-footer" in html_content:
            return html_content

        header = (
            '<div class="sa-corp-header">'
            '<span><strong>SuperAgente IA Pro</strong> · Informe Ejecutivo</span>'
            '<span style="float:right;">Documento Confidencial</span>'
            "</div>"
        )
        footer = (
            '<div class="sa-corp-footer">'
            '<span>Generado por SuperAgente IA Pro</span>'
            '<span style="float:right;">Página <span class="sa-page-number"></span></span>'
            "</div>"
        )

        if "<body" in html_content.lower():
            html_content = re.sub(r"(?i)(<body[^>]*>)", r"\1" + header, html_content, count=1)
            html_content = re.sub(r"(?i)</body>", footer + r"</body>", html_content, count=1)
            return html_content

        return header + html_content + footer

